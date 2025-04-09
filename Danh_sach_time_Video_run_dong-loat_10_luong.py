import os
import subprocess
import openpyxl
from datetime import datetime
from tqdm import tqdm
import re
import threading
import queue
import logging
import time
from colorama import Fore, Style, init

# Khởi tạo colorama để hỗ trợ màu sắc trong terminal
init(autoreset=True)

# Đường dẫn file Log.txt
log_file = 'Log.txt'

# Hàm ghi log mới vào đầu file
def write_new_logs(new_logs):
    with open(log_file, 'r+', encoding='utf-8') as f:
        # Đọc nội dung cũ
        old_logs = f.read()
        # Di chuyển con trỏ về đầu file
        f.seek(0)
        # Ghi log mới vào đầu file
        f.write(new_logs + "\n" + old_logs)

# Đo thời gian bắt đầu và hiển thị
start_time = time.time()
start_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
print(f"{Fore.CYAN}Bắt đầu chạy chương trình vào: {start_time_str}{Style.RESET_ALL}")

# Danh sách thư mục cần quét
work_queue = queue.Queue()
drive_paths = [
    (r"G:\.shortcut-targets-by-id\13cEuVXiiAR1ftfXwFQGv9VaBo4JiNv5y\2D\Video Thô\Video Thô", "2D"),
    (r"G:\.shortcut-targets-by-id\13cEuVXiiAR1ftfXwFQGv9VaBo4JiNv5y\2D\Video Thô\Video Thô Thay NV", "2D TNV GỐC"),
    (r"G:\Shared drives\BT Group 2D\Dự án Horror Stories Animated\Video Thô", "Horror"),
    (r"G:\Shared drives\BT Group 2D\Hậu kỳ\Video FINAL\1_Video GỐC\Tiếng Anh", "Hậu Kỳ Gốc EN"),
    (r"G:\Shared drives\BT Group 2D\Hậu kỳ\Video FINAL\Không phận sự MIỄN VÀO", "OS Team cũ"),
    (r"G:\.shortcut-targets-by-id\13cEuVXiiAR1ftfXwFQGv9VaBo4JiNv5y\2D\Project animation\2025\OS\Video Thô", "Diễn OS")
]

for path in drive_paths:
    work_queue.put(path)

# Đường dẫn file Excel
output_file = "Danh sách thời lượng Video thô.xlsx"

# Kiểm tra xem file Excel đã tồn tại chưa
if os.path.exists(output_file):
    wb = openpyxl.load_workbook(output_file)
else:
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

# Biểu thức chính quy để kiểm tra tên file và trích xuất số tập
pattern = re.compile(r'^(Tập|TAP|TẠP|tập|Tạp|TẬP)\s*(\d+)', re.IGNORECASE)

# Khóa để tránh xung đột khi ghi file Excel
excel_lock = threading.Lock()

# Giới hạn số luồng tối đa
MAX_THREADS = 12
SPLIT_THRESHOLD = 100  # Ngưỡng chia luồng xử lý file

# Biến để đếm tổng số file và lưu kết quả từ các đường dẫn
total_files = 0
results = {}  # Lưu dữ liệu theo sheet_name

def extract_episode_number(filename):
    """Trích xuất số tập từ tên file, nếu không có trả về -1."""
    match = pattern.match(filename)
    return int(match.group(2)) if match else -1

def process_files(sheet_name, video_files, result_list):
    """Chia nhỏ danh sách file và thu thập dữ liệu vào result_list."""
    if len(video_files) > SPLIT_THRESHOLD:
        num_threads = min(MAX_THREADS, len(video_files))  # Đảm bảo không vượt quá số file
    else:
        num_threads = 1
    
    def worker(files):
        progress_bar = tqdm(files, desc=f"Đang kiểm tra video ({sheet_name})", unit="file", colour='cyan')
        for i, video_path in enumerate(progress_bar):
            try:
                file = os.path.basename(video_path)
                creation_time = os.path.getctime(video_path)
                creation_date = datetime.fromtimestamp(creation_time).strftime("%Y-%m-%d")
                
                cmd = f'ffmpeg -i "{video_path}"'
                result = subprocess.run(cmd, shell=True, stderr=subprocess.PIPE, stdout=subprocess.PIPE, text=True, encoding="utf-8", errors="ignore")
                
                total_seconds = 0
                duration_formatted = "00:00"
                video_size = "N/A"
                
                for line in result.stderr.split("\n"):
                    if "Duration" in line:
                        duration_str = line.split(",")[0].split("Duration: ")[1].strip()
                        h, m, s = map(float, duration_str.replace(":", " ").split())
                        total_seconds = int(h * 3600 + m * 60 + s)
                        duration_formatted = f"{total_seconds // 60:02}:{total_seconds % 60:02}"
                    if "Video:" in line:
                        resolution_match = re.search(r"(\d{3,4}x\d{3,4})", line)
                        if resolution_match:
                            video_size = resolution_match.group(1)
                
                result_list.append([file, creation_date, total_seconds, duration_formatted, video_size, video_path])
            except Exception as e:
                logging.error(f"❌ Lỗi khi xử lý tệp {video_path}: {e}")
            if i > len(files) * 0.8:
                progress_bar.colour = 'blue'
            if i > len(files) * 0.9:
                progress_bar.colour = 'green'
    
    chunk_size = len(video_files) // num_threads if num_threads > 1 else len(video_files)
    threads = []
    for i in range(num_threads):
        start = i * chunk_size
        end = None if i == num_threads - 1 else (i + 1) * chunk_size
        thread = threading.Thread(target=worker, args=(video_files[start:end],))
        threads.append(thread)
        thread.start()
    
    for thread in threads:
        thread.join()

def process_drive():
    global total_files
    while not work_queue.empty():
        drive, sheet_name = work_queue.get()
        try:
            if not os.path.exists(drive):
                logging.warning(f"⚠️ Đường dẫn không tồn tại: {drive}")
                print(f"{Fore.YELLOW}⚠️ Đường dẫn không tồn tại: {drive}{Style.RESET_ALL}")
                continue
            
            video_files = [os.path.join(root, file) for root, _, files in os.walk(drive) for file in files if file.lower().endswith(".mp4") and pattern.match(file)]
            
            folder_name = os.path.basename(drive.rstrip("\\/"))
            print(f"{Fore.YELLOW}Đã tìm thấy {len(video_files)} video trong {folder_name}\n{Style.RESET_ALL}")
            total_files += len(video_files)
            
            result_list = []
            process_files(sheet_name, video_files, result_list)
            
            # Sắp xếp giảm dần theo số tập, nếu số tập bằng nhau thì theo ngày tạo tăng dần
            result_list.sort(key=lambda x: (extract_episode_number(x[0]), x[1]), reverse=True)
            
            # Lưu kết quả vào dictionary
            results[sheet_name] = result_list
            
        except Exception as e:
            logging.error(f"❌ Lỗi khi xử lý đường dẫn {drive}: {e}")
            print(f"{Fore.RED}❌ Lỗi khi xử lý đường dẫn {drive}: {e}{Style.RESET_ALL}")

# Chạy các luồng để xử lý song song
threads = []
for _ in range(min(MAX_THREADS, len(drive_paths))):  # Không vượt quá số đường dẫn
    thread = threading.Thread(target=process_drive)
    threads.append(thread)
    thread.start()

for thread in threads:
    thread.join()

# Ghi dữ liệu vào workbook theo thứ tự trong drive_paths
for drive, sheet_name in drive_paths:
    if sheet_name in results:  # Chỉ ghi nếu đường dẫn đã được xử lý thành công
        with excel_lock:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.delete_rows(2, ws.max_row)
            else:
                ws = wb.create_sheet(sheet_name)
                ws.append(["Tên file", "Ngày tạo file", "Thời lượng (giây)", "Thời lượng (mm:ss)", "Kích thước video", "Đường dẫn file"])
            
            result_list = results[sheet_name]
            seen_files = {}
            for row in result_list:
                file_name = row[0]
                if file_name in seen_files:
                    ws.append([row[0], row[1], row[2], row[3], row[4], row[5]])
                    seen_files[file_name] += 1
                else:
                    ws.append([row[0], row[1], row[2], row[3], row[4], ""])
                    seen_files[file_name] = 1

# Đo thời gian kết thúc và tính tổng thời gian chạy
end_time = time.time()
end_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
total_time = end_time - start_time
total_minutes = int(total_time // 60)
total_seconds = int(total_time % 60)
total_time_formatted = f"{total_minutes:02}:{total_seconds:02}"

print(f"{Fore.GREEN}✅ Done: {end_time_str}{Style.RESET_ALL}")
print(f"{Fore.CYAN}Tổng thời gian chạy: {total_time_formatted}{Style.RESET_ALL}")

# Tạo nội dung log mới
new_logs = f"Bắt đầu chạy chương trình vào: {start_time_str}\n"
for drive, sheet_name in drive_paths:
    if sheet_name in results:
        new_logs += f"Thư mục: {drive} - Số lượng video: {len(results[sheet_name])}\n"
new_logs += f"Thời gian hoàn thành: {end_time_str}\n"
new_logs += f"Tổng thời gian chạy: {total_time_formatted}\n"
new_logs += f"Tổng số file: {total_files}\n"
new_logs += f"✅ Đã lưu danh sách vào {output_file}\n"

# Ghi log mới vào đầu file
write_new_logs(new_logs)

wb.save(output_file)
print(f"{Fore.GREEN}✅ Đã lưu danh sách vào {output_file}{Style.RESET_ALL}")